from io import BytesIO

from openpyxl import Workbook
from openpyxl.utils import get_column_letter


def generate_excel(report_type: str, report_data: dict) -> BytesIO:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = report_data.get("title", "Report")[:31]

    if report_type == "dashboard":
        summary = report_data.get("summary", {})
        rows = [
            ["Metric", "Value"],
            ["Total Candidates", summary.get("total_candidates", 0)],
            ["Total Jobs", summary.get("total_jobs", 0)],
            ["Total Applications", summary.get("total_applications", 0)],
            ["Total Screenings", summary.get("total_screenings", 0)],
        ]
    else:
        rows = report_data.get("rows", [])
        if rows:
            headers = list(rows[0].keys())
            rows = [headers] + [list(item.values()) for item in rows]
        else:
            rows = [["No data available"]]

    for row_index, row in enumerate(rows, start=1):
        for col_index, value in enumerate(row, start=1):
            sheet.cell(row=row_index, column=col_index, value=value)

    for col_index in range(1, sheet.max_column + 1):
        column_letter = get_column_letter(col_index)
        sheet.column_dimensions[column_letter].auto_size = True

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer
